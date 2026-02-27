import utils.FileTools as file
from openpyxl import load_workbook

PATH_TRANSLATE_ROOT = "/Users/user/Downloads/translate/"
TRANSLATE_STRING_FILE = PATH_TRANSLATE_ROOT + "strings.xml"
STRING_TAG = '<string'
STRING_ID_TAG = '"'
STRING_VALUE_TAG1 = '>'
STRING_VALUE_TAG2 = '<'
TRANSLATE_FILE = PATH_TRANSLATE_ROOT + "【翻译】oppo海外8.1.0-全语言.xlsx"
TRANSLATE_OLD_FILE = PATH_TRANSLATE_ROOT + "【翻译】oppo海外8.1.0-全语言.xlsx"


DNY_TRANSLATE_FILE = PATH_TRANSLATE_ROOT + "【翻译】oppo海外8.1.0-全语言.xlsx"
DNY_TRANSLATE_OLD_FILE = PATH_TRANSLATE_ROOT + "【翻译】oppo海外8.1.0-全语言.xlsx"

ROW = 4
COLUMN = 4
INDEX_SHEET = 4

ids_same = []
values_same = []

ids_check = []
values_check = []

#内置安规问题，有些字符串不需要导入，否则有安规问题
# useless_key =["reminder_privacy_content","eu_recommend_setting_close_desc","close_online_service_desc","android_base_app_update_message"]

def start_check():
    string_result = file.read_file(TRANSLATE_STRING_FILE)
    index = 0
    ids = []
    values = []
    while True:
        index = string_result.find(STRING_TAG, index)
        if index > 0:
            index = string_result.index(STRING_ID_TAG, index)
            if index >= 0:
                index2 = string_result.index(STRING_ID_TAG, index + 1)
                if index2 >= 0:
                    id = string_result[index + 1:index2]
                    # if id in useless_key:
                    #     print("过滤掉了安规字符串"+id)
                    #     continue
                    ids.append(id)
                    # print(id)
                    index = index2 + 1
                    index = string_result.index(STRING_VALUE_TAG1, index)
                    if index >= 0:
                        index2 = string_result.index(
                            STRING_VALUE_TAG2, index + 1)
                        value = string_result[index + 1:index2]

                        values.append(value)
                        # print(value)
                        index = index2 + 1
        else:
            break
    for index in range(0, len(ids)):
        if values[index].find("@string/") >= 0:
            ids_same.append(ids[index])
            values_same.append(values[index])
            # print(link_string(ids[index], values[index]))
        else:
            ids_check.append(ids[index])
            values_check.append(values[index])


# 从excel里面读取的资源ID
excel_ids = []
excel_languages = dict()
# 存储合并后的翻译结果
merged_languages = dict()


def link_string(id, value):
    return f'\t<string name="{id}">{value}</string>'


def link_string_none_translatable(id, value):
    return f'\t<string name="{id}" translatable="false">{value}</string>'


def write_string_xml(path, v):
    result = '<?xml version="1.0" encoding="utf-8"?>\n<resources>\n' + \
             "\n".join(v) + '\n</resources>'
    file.write_file(path, result)


def read_excel_file(file_path, is_first_file=False):
    global excel_ids
    if is_first_file:
        excel_ids = []  # 如果是第一个文件，清空ID列表
        
    wb = load_workbook(file_path)
    sheet1 = wb['客户端识别码']  # 通过已知表名获取sheet
    
    max_value = 0
    for index in range(ROW, 2000):  # 从第四行开始读取
        value = sheet1.cell(index, 2).value
        max_value += 1
        if value is None:
            break
    print(f"max_value:{max_value}")
    
    # 记录当前文件中的ID
    current_file_ids = []
    for idIndex in range(ROW, 2000):
        value = sheet1.cell(idIndex, 2).value
        if value is None:
            break
        if value in excel_ids and not value.endswith("_plural"):
            current_file_ids.append(value + "_plural")
        else:
            current_file_ids.append(value)
            excel_ids.append(value)
    
    # 读取当前文件的翻译内容
    current_file_languages = {}
    for index in range(COLUMN, 1000):
        title = sheet1.cell(1, index).value
        if title is None:
            break
        values_language = []
        for index2 in range(ROW, max_value + ROW):
            value = str(sheet1.cell(index2, index).value)
            value = value.replace('&', '&amp;')
            value = value.replace('\'', '\\\'')
            value = value.replace('\\\\\'', '\\\'')
            values_language.append(value)
        current_file_languages[title] = values_language
    
    # 合并翻译
    for lang, translations in current_file_languages.items():
        if lang not in merged_languages:
            merged_languages[lang] = {}
        
        # 保存当前文件的翻译到合并结果中
        for i, str_id in enumerate(current_file_ids):
            if i < len(translations) and translations[i] and translations[i] != "None":
                merged_languages[lang][str_id] = translations[i]


def generate_string_xmls():
    for ll in merged_languages.keys():
        v = []
        # 为每个ID生成翻译
        for i in range(0, len(ids_check)):
            str_id = ids_check[i]
            # 检查是否有该ID的翻译
            if str_id in merged_languages[ll]:
                value = merged_languages[ll][str_id]
                if ll == "us" or len(value) > 0:
                    v.append(link_string(str_id, value))
            
        if ll == "values-en-rUS":
            # 添加不需要翻译的字符串
            for i in range(0, len(ids_same)):
                v.append(link_string(ids_same[i], values_same[i]))
                v.append("\n")
            
            # 添加未找到翻译的字符串（使用原始值）
            for i in range(0, len(ids_check)):
                str_id = ids_check[i]
                if str_id not in merged_languages[ll]:
                    v.append(link_string_none_translatable(str_id, values_check[i]))
            
            write_string_xml(PATH_TRANSLATE_ROOT + "res//values//strings.xml", v)
        else:
            write_string_xml(PATH_TRANSLATE_ROOT + f"res//{ll}//strings.xml", v)
            if ll == "values-b+sr+Latn":
                write_string_xml(PATH_TRANSLATE_ROOT + f"res//values-b+bs+BA//strings.xml", v)


if __name__ == '__main__':
    start_check()
    
    # 读取第一个翻译文件
    read_excel_file(TRANSLATE_OLD_FILE, True)
    
    # 读取第二个翻译文件
    read_excel_file(DNY_TRANSLATE_OLD_FILE, False)
    
    # 生成最终的strings.xml文件
    generate_string_xmls()
